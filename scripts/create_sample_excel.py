#!/usr/bin/env python3
"""
Create a sample Excel (.xlsx) file with the scheme table format (header + one example row).
Use this as a template: fill in your schemes and run excel_to_schemes_sql.py on it.

Usage:
  python create_sample_excel.py [output.xlsx]
"""

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

HEADER = [
    "Name",
    "Type",
    "Benefit Summary",
    "Key Benefit Display",
    "States",
    "Required Documents",
    "Estimated Timeline",
    "Business Types",
    "Industries",
    "Turnover Max",
    "Turnover Min",
    "Company Age Max",
    "Company Age Min",
    "Funding Types",
]

# One example row (PMEGP-like)
EXAMPLE_ROW = [
    "PMEGP",
    "subsidy",
    "Credit-linked capital subsidy for micro enterprises. Margin money assistance.",
    "Up to 25% margin money subsidy",
    "*",
    "Business plan; Identity proof; Address proof; Bank statement",
    "4–6 weeks from application",
    "micro, small",
    "",
    "100",
    "",
    "3",
    "",
    "subsidy",
]


def main():
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("schemes_template.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Schemes"
    # Header row (bold)
    for c, text in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=c, value=text)
        cell.font = Font(bold=True)
    # Example row
    for c, text in enumerate(EXAMPLE_ROW, start=1):
        ws.cell(row=2, column=c, value=text)
    wb.save(out_path)
    print(f"Created {out_path.absolute()}")
    print("Add more rows with your scheme data, then run:")
    print(f"  python excel_to_schemes_sql.py {out_path.name} > seed_schemes.sql")


if __name__ == "__main__":
    main()
