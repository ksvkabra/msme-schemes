#!/usr/bin/env python3
"""
Read an Excel (.xlsx) file with one row per scheme and generate INSERT SQL for public.schemes.
First row must be the header. See README_SCHEME_IMPORT.md for the expected column names.

Usage:
  python excel_to_schemes_sql.py path/to/schemes.xlsx [> output.sql]
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Expected header names (case-insensitive) -> key for row dict
HEADER_KEYS = {
    "name": "name",
    "type": "type",
    "benefit summary": "benefit_summary",
    "key benefit display": "key_benefit_display",
    "states": "states",
    "required documents": "required_documents",
    "estimated timeline": "estimated_timeline",
    "business types": "business_types",
    "industries": "industries",
    "turnover max": "turnover_max",
    "turnover min": "turnover_min",
    "company age max": "company_age_max",
    "company age min": "company_age_min",
    "funding types": "funding_types",
}


def cell_value(cell):
    """Get string value from a cell (number/date converted to string)."""
    if cell is None or cell.value is None:
        return ""
    v = cell.value
    if isinstance(v, (int, float)):
        return str(v).strip()
    return (str(v) or "").strip()


def parse_header(row):
    """Map column index (1-based in openpyxl) -> key. Row is iterable of cells."""
    col_map = {}
    for i, cell in enumerate(row):
        raw = cell_value(cell).lower().strip()
        if raw in HEADER_KEYS:
            col_map[i] = HEADER_KEYS[raw]
    return col_map


def row_to_dict(row, col_map):
    """Build a dict of key -> value for this row."""
    out = {}
    for i, cell in enumerate(row):
        if i in col_map:
            out[col_map[i]] = cell_value(cell)
    return out


def parse_list(s, sep=","):
    """Split by sep and return non-empty stripped items."""
    if not s or not str(s).strip():
        return []
    return [x.strip() for x in str(s).split(sep) if x.strip()]


def parse_documents(s):
    """Required documents: semicolon-separated."""
    return parse_list(s, ";")


def parse_number(s):
    """Return int or float from string, else None."""
    if s is None or (isinstance(s, str) and not s.strip()):
        return None
    s = str(s).strip().replace(",", "")
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return None


def build_eligibility_rules(data):
    """Build eligibility_rules JSON object from row data."""
    rules = {}
    bt = parse_list(data.get("business_types") or "")
    if bt:
        rules["business_types"] = [x.lower() for x in bt]
    ind = parse_list(data.get("industries") or "")
    if ind:
        rules["industries"] = ind
    states = (data.get("states") or "").strip() if isinstance(data.get("states"), str) else str(data.get("states") or "").strip()
    if states and states != "*":
        rules["states"] = parse_list(states)
    elif states == "*":
        rules["states"] = ["*"]
    tmax = parse_number(data.get("turnover_max"))
    if tmax is not None:
        rules["turnover_max"] = tmax
    tmin = parse_number(data.get("turnover_min"))
    if tmin is not None:
        rules["turnover_min"] = tmin
    camax = parse_number(data.get("company_age_max"))
    if camax is not None:
        rules["company_age_max_years"] = int(camax)
    camin = parse_number(data.get("company_age_min"))
    if camin is not None:
        rules["company_age_min_years"] = int(camin)
    ft = parse_list(data.get("funding_types") or "")
    if ft:
        rules["funding_types"] = [x.lower() for x in ft]
    return rules


def sql_escape(s):
    """Escape single quotes for PostgreSQL string literal."""
    if s is None:
        return ""
    return (str(s) or "").replace("\\", "\\\\").replace("'", "''")


def sql_string(s):
    """Return SQL literal for a string (including NULL for empty)."""
    if s is None or (isinstance(s, str) and not s.strip()):
        return "null"
    return "'" + sql_escape(s.strip()) + "'"


def sql_string_array(arr):
    """PostgreSQL array literal for text[]."""
    if not arr:
        return "'{}'"
    parts = ["'" + sql_escape(x) + "'" for x in arr]
    return "array[" + ", ".join(parts) + "]"


def main():
    if len(sys.argv) < 2:
        print("Usage: python excel_to_schemes_sql.py <path/to/schemes.xlsx>", file=sys.stderr)
        sys.exit(1)
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows())
    wb.close()
    if len(rows) < 2:
        print("Sheet must have header row and at least one data row.", file=sys.stderr)
        sys.exit(1)
    col_map = parse_header(rows[0])
    if "name" not in col_map or "type" not in col_map or "benefit_summary" not in col_map:
        print("Sheet must have columns: Name, Type, Benefit Summary", file=sys.stderr)
        sys.exit(1)

    inserts = []
    for row in rows[1:]:
        data = row_to_dict(row, col_map)
        name = (data.get("name") or "").strip()
        if not name:
            continue
        stype = (data.get("type") or "").strip().lower()
        if stype not in ("loan", "subsidy", "grant"):
            print(f"Skipping row (invalid type '{stype}'): {name}", file=sys.stderr)
            continue
        benefit_summary = (data.get("benefit_summary") or "").strip()
        if not benefit_summary:
            print(f"Skipping row (empty Benefit Summary): {name}", file=sys.stderr)
            continue

        eligibility = build_eligibility_rules(data)
        eligibility_json = json.dumps(eligibility, ensure_ascii=False)
        key_benefit = (data.get("key_benefit_display") or "").strip() or None
        states_raw = (data.get("states") or "").strip()
        if states_raw == "*" or not states_raw:
            states_applicable = "null"
        else:
            states_list = parse_list(states_raw)
            states_applicable = sql_string_array(states_list)
        docs = parse_documents(data.get("required_documents") or "")
        timeline = (data.get("estimated_timeline") or "").strip() or None

        values = (
            sql_string(name),
            sql_string(stype),
            "'" + sql_escape(eligibility_json) + "'::jsonb",
            sql_string(benefit_summary),
            states_applicable,
            sql_string(key_benefit),
            sql_string_array(docs) if docs else "'{}'",
            sql_string(timeline),
        )
        inserts.append(
            "insert into public.schemes (name, type, eligibility_rules, benefit_summary, states_applicable, key_benefit_display, required_documents, estimated_timeline)\n"
            f"values ({', '.join(values)});"
        )

    if not inserts:
        print("No valid rows to insert.", file=sys.stderr)
        sys.exit(1)

    print("-- Generated from", path.name)
    print("-- Run in Supabase SQL Editor or psql\n")
    for stmt in inserts:
        print(stmt)
        print()


if __name__ == "__main__":
    main()
